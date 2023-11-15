import openpyxl
#엑셀만들기
wb=openpyxl.Workbook()
#엑셀 워크시트만들기
ws=wb.create_sheet('오징어게임')
#데이터추가하기
ws['A1']='참가번호'
ws['B1']='성명'
ws['A2']=1
ws['B2']='오일남'
#엑셀 저장하기
wb.save(r'C:\python\test\참가자_data.xlsx')